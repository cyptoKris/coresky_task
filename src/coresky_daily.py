import json
import os
import random
import time

import requests
from Crypto.Cipher import AES
from Crypto.Util.Padding import pad
import base64
import ddddocr

from eth_account import Account
from eth_account.messages import encode_defunct
from openpyxl.reader.excel import load_workbook


def wr(_id):
    with open(r"../完成的账户id.txt", "a+") as f:
        f.write(_id + "\n")


def rd():
    try:
        with open(r"../完成的账户id.txt", "r+") as f:
            return [i.replace("\n", "") for i in f.readlines()]
    except FileNotFoundError:
        return []


def safe_request(method, url, session, max_retries=3, delay=2, **kwargs):
    for attempt in range(1, max_retries + 1):
        try:
            if method.lower() == 'get':
                response = session.get(url, **kwargs)
            elif method.lower() == 'post':
                response = session.post(url, **kwargs)
            else:
                raise ValueError("只支持 GET 和 POST 方法")
            response.raise_for_status()
            return response
        except Exception as e:
            print(f"⚠️ 请求失败 (第 {attempt} 次): {e}")
            if attempt == max_retries:
                raise
            time.sleep(delay)


class Coresky:
    def __init__(self, id, okx_address, okx_private_key, proxy):
        self.id = id
        self.okx_address = okx_address
        self.okx_private_key = okx_private_key
        self.session = requests.Session()
        self.proxy = proxy
        self.session.headers.update({
            'accept': 'application/json, text/plain, */*',
            'accept-language': 'zh-CN,zh;q=0.9',
            'cache-control': 'no-cache',
            'content-type': 'application/json',
            'hearder_gray_set': '0',
            'origin': 'https://www.coresky.com',
            'pragma': 'no-cache',
            'priority': 'u=1, i',
            'referer': 'https://www.coresky.com/tasks-rewards',
            'sec-ch-ua': '"Not)A;Brand";v="99", "Google Chrome";v="127", "Chromium";v="127"',
            'sec-ch-ua-mobile': '?0',
            'sec-ch-ua-platform': '"Windows"',
            'sec-fetch-dest': 'empty',
            'sec-fetch-mode': 'cors',
            'sec-fetch-site': 'same-origin',
            'token': '',
            'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/127.0.0.0 Safari/537.36',
        })
        try:
            ipinfo = safe_request('get', "http://ipinfo.io", session=self.session, proxies=self.proxy)
            print("当前代理：", ipinfo.json())
        except Exception as e:
            print("代理错误", e)

    def checksum_address(self, address):
        return address

    def okx_sign_message(self, message, private_key):
        account = Account.from_key(private_key)
        message_hash = encode_defunct(text=message)
        signature = account.sign_message(message_hash)
        return signature.signature.hex()

    @staticmethod
    def get_message(address):
        return (f"Welcome to CoreSky!\n\nClick to sign in and accept the CoreSky Terms of Service.\n\nThis request "
                f"will not trigger a blockchain transaction or cost any gas fees.\n\nYour authentication status "
                f"will reset after 24 hours.\n\nWallet address:\n\n{address}")

    def login(self, address, signature):
        json_data = {
            'address': address,
            'signature': signature,
            'refCode': '7f9ch5mb',
            'projectId': '',
        }
        response = safe_request('post', 'https://www.coresky.com/api/user/login', session=self.session, json=json_data, proxies=self.proxy)
        if response.json().get("debug").get("token") is None:
            raise Exception("接口返回token错误")
        token = response.json().get("debug").get("token")
        print(f"账户{self.id} 登录成功 token: {token}")
        self.session.headers.update({'hearder_gray_set': '0', 'token': token})
        safe_request('get', 'https://www.coresky.com/api/user/message/list', session=self.session, proxies=self.proxy)

    @staticmethod
    def solve_slider(jigsaw_path, background_path):
        ocr = ddddocr.DdddOcr(det=False, ocr=False)
        with open(jigsaw_path, "rb") as f:
            target_bytes = f.read()
        with open(background_path, "rb") as f:
            background_bytes = f.read()
        try:
            res = ocr.slide_match(target_bytes, background_bytes)
            print(res)
            target = res['target']
            return target[0], target[1]
        except Exception as e:
            print("None", e)
            return None

    def sign_generation(self, max_retries=5):
        for attempt in range(1, max_retries + 1):
            try:
                print(f"账户{self.id} 第 {attempt} 次尝试获取验证码并验证...")

                # 1. 获取验证码
                json_data = {'captchaType': 'blockPuzzle'}
                response1 = safe_request('post', 'https://www.coresky.com/api/captcha/get', session=self.session,
                                         json=json_data, proxies=self.proxy)
                hua, yuan = response1.json()['debug']['jigsawImageBase64'], response1.json()['debug'][
                    'originalImageBase64']
                hua = hua.split(",")[1] if hua.startswith("data:image") else hua
                yuan = yuan.split(",")[1] if yuan.startswith("data:image") else yuan

                with open("jigsaw.png", "wb") as f:
                    f.write(base64.b64decode(hua))
                with open("original.png", "wb") as f:
                    f.write(base64.b64decode(yuan))
                print("✅ 验证码图片已保存")

                # 2. 本地识别滑块位置
                result = self.solve_slider("jigsaw.png", "original.png")
                if not result:
                    raise Exception("滑块识别失败")
                x, y = result
                data = {"x": x, "y": 5}
                e = json.dumps(data, separators=(',', ':'))
                d_value = response1.json()['debug']["secretKey"]
                time.sleep(random.uniform(3.5, 4.5))
                encrypted_result = self.post(e, d_value)

                # 3. 提交验证码校验
                json_data = {
                    'captchaType': 'blockPuzzle',
                    'pointJson': encrypted_result,
                    'token': response1.json()['debug']["token"],
                }
                response2 = safe_request('post', 'https://www.coresky.com/api/captcha/check', session=self.session,
                                         json=json_data, proxies=self.proxy)
                if response2.json().get('code') != 200:
                    print(f"⚠️ 账户{self.id} 验证码验证失败，将重新尝试")
                    time.sleep(random.uniform(1, 2))
                    continue  # 失败就重新获取验证码

                # 4. 验证码通过后，提交签到签名
                print(f"✅ 账户{self.id} 验证码通过，准备提交签到")
                json_data = {
                    'responseToken': self.post(response2.json()['debug']["token"] + "---" + e, d_value),
                }
                response3 = safe_request('post', 'https://www.coresky.com/api/taskwall/meme/sign', session=self.session,
                                         json=json_data, proxies=self.proxy)
                os.remove("jigsaw.png")
                os.remove("original.png")

                if response3.json().get('code') == 200:
                    print(f"🎉 {self.id} 签到成功")
                    wr(self.id)
                    return
                else:
                    raise Exception("提交签到失败")

            except Exception as e:
                print(f"❌ 第 {attempt} 次验证码流程失败：{e}")
                time.sleep(random.uniform(1.5, 2.5))

        raise Exception(f"账户{self.id} 多次验证码尝试均失败，终止任务")

    @staticmethod
    def post(e, tt):
        key_bytes = tt.encode('utf-8')
        plaintext_bytes = e.encode('utf-8')
        padded_bytes = pad(plaintext_bytes, AES.block_size, style='pkcs7')
        cipher = AES.new(key_bytes, AES.MODE_ECB)
        encrypted_bytes = cipher.encrypt(padded_bytes)
        return base64.b64encode(encrypted_bytes).decode('utf-8')

    def check_task(self):
        response = safe_request('get', 'https://www.coresky.com/api/taskwall/meme/tasks', session=self.session, proxies=self.proxy)
        if response.json().get("debug") is None:
            raise Exception("获取任务失败")
        for task in response.json().get("debug"):
            if "Daily Check-in" == task.get("taskName"):
                if task.get('taskStatus') != 2:
                    self.sign_generation()
                else:
                    print(f"{self.id} 不需要签到")

    def vote(self):
        response = safe_request('post', 'https://www.coresky.com/api/user/token', session=self.session, json={}, proxies=self.proxy)
        if response.json().get("message") != "success":
            raise Exception("获取个人信息失败")
        score = response.json().get('debug').get("score")
        print(f"账户{self.id} 当前分数: {score}")
        if score > 0:
            json_data = {
                'projectId': 51,
                'voteNum': score,
            }
            safe_request('post', 'https://www.coresky.com/api/taskwall/meme/vote', session=self.session, json=json_data, proxies=self.proxy)
            print(f"账户{self.id} 投票第一名{score}分数成功")
        else:
            print(f"账户{self.id} 分数不足无法投票")

    @staticmethod
    def read_accounts_from_excel(file_path):
        wb = load_workbook(file_path, read_only=True)
        ws = wb.active
        accounts = []
        for i, row in enumerate(ws.iter_rows(min_row=2), start=1):
            address = row[1].value
            private_key = row[2].value
            if not address or not private_key:
                continue
            accounts.append({
                "id": address.strip()[:6],
                "address": address.strip(),
                "private_key": private_key.strip()
            })
        return accounts

    def run(self):
        try:
            message = self.get_message(self.checksum_address(self.okx_address))
            signature = self.okx_sign_message(message, self.okx_private_key)
            self.login(self.checksum_address(self.okx_address), signature)
            self.check_task()
            self.vote()
            print(f"账号{self.id} 任务完成")
        except Exception as e:
            print(f"账号{self.id} 任务失败: {str(e)}")


if __name__ == '__main__':
    accounts = Coresky.read_accounts_from_excel("../模版.xlsx")
    proxy = {

    }

    MAX_RETRIES = 3
    for attempt in range(1, MAX_RETRIES + 1):
        print(f"\n🔁 第 {attempt} 轮任务执行开始")
        completed_ids = rd()
        remaining_accounts = [acc for acc in accounts if acc["id"] not in completed_ids]
        if not remaining_accounts:
            print("✅ 所有账号任务已完成，无需重复执行")
            break

        for account in remaining_accounts:
            coresky = Coresky(
                id=account["id"],
                okx_address=account["address"],
                okx_private_key=account["private_key"],
                proxy=proxy
            )
            try:
                coresky.run()
            except Exception as e:
                print(f"❌ 账号 {account['id']} 执行出错: {e}")
            time.sleep(random.uniform(1, 3))

        print(f"✅ 第 {attempt} 轮任务执行完毕")

    if all(acc["id"] in rd() for acc in accounts):
        print("\n🎉 所有账号任务已全部成功完成！")
    else:
        print("\n⚠️ 多轮尝试后，仍有部分账号失败，请检查日志或手动处理。")
